"""
Excel导出器 — 支持多Sheet（按日期分组）+ 单据头部信息
"""
import os
import re
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR, EXCEL_HEADERS


# 需要纯数字的列
_NUMERIC_ONLY_COLUMNS = {"数量", "合计", "备注"}


def _extract_number(text: str) -> str:
    """从文本中提取纯数字，丢弃所有非数字字符（保留小数点）"""
    text = str(text or "").strip()
    if not text:
        return ""
    m = re.search(r"(\d+\.?\d*)", text)
    return m.group(1) if m else text


# === 样式常量 ===
TITLE_FONT = Font(name="微软雅黑", size=16, bold=True)
INFO_FONT = Font(name="微软雅黑", size=10)
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="微软雅黑", size=10)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COL_WIDTHS = {
    "序号": 8, "名称": 24, "单位": 8,
    "购定价（含税）": 16, "税率": 8, "网上询价": 14,
    "数量": 10, "合计": 14, "备注": 20,
}

# 写入格式常量
HEADER_ROW = 5   # 表头在第5行
DATA_START = 6   # 数据从第6行开始


def _write_title_rows(ws, header_info: dict):
    """向工作表写入标题行（公司名 + 收货单位/日期）"""
    num_cols = len(EXCEL_HEADERS)

    company = header_info.get("company", "") if header_info else ""
    consignee = header_info.get("consignee", "") if header_info else ""
    date_str = header_info.get("date", "") if header_info else ""

    # 行1: 公司名称居中加粗
    if company:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        cell = ws.cell(row=1, column=1, value=company)
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 行2: 空
    ws.row_dimensions[2].height = 6

    # 行3: 收货单位(左) + 送货日期(右)
    if consignee or date_str:
        left_text = f"收货单位：{consignee}" if consignee else ""
        right_text = f"送货日期：{date_str}" if date_str else ""

        if left_text:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(1, num_cols // 2))
            c = ws.cell(row=3, column=1, value=left_text)
            c.font = INFO_FONT
            c.alignment = Alignment(horizontal="left", vertical="center")

        if right_text:
            ws.merge_cells(start_row=3, start_column=num_cols // 2 + 1, end_row=3, end_column=num_cols)
            c = ws.cell(row=3, column=num_cols // 2 + 1, value=right_text)
            c.font = INFO_FONT
            c.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 6


def _write_sheet(ws, rows: list[dict], header_info: dict = None):
    """向工作表写入标题行、表头和数据行"""
    # 标题行
    _write_title_rows(ws, header_info)

    # 表头（白色背景）
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    ws.row_dimensions[HEADER_ROW].height = 28

    # 数据
    for row_idx, row_data in enumerate(rows, DATA_START):
        for col_idx, header in enumerate(EXCEL_HEADERS, 1):
            cell_data = row_data.get(header, "")
            value = cell_data.get("value", cell_data) if isinstance(cell_data, dict) else str(cell_data or "")
            # 数量/合计/备注列只保留纯数字
            if header in _NUMERIC_ONLY_COLUMNS:
                value = _extract_number(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if header in ("名称", "备注"):
                cell.alignment = LEFT_ALIGNMENT
            else:
                cell.alignment = CENTER_ALIGNMENT

    # 列宽
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 12)

    # 冻结表头下方
    ws.freeze_panes = f"A{DATA_START}"

    # 打印设置
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def export_excel_single_sheet(rows: list[dict], filename: str,
                               sheet_name: str = "送货单",
                               header_info: dict = None) -> str:
    """导出为单Sheet Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_sheet(ws, rows, header_info)

    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath


def export_excel_multi_sheet(date_groups: dict[str, list[dict]], filename: str,
                              header_info: dict = None) -> str:
    """按日期分Sheet导出Excel"""
    wb = Workbook()
    wb.remove(wb.active)

    first = True
    for date_str, rows in date_groups.items():
        sheet_name = date_str[:31]
        if first:
            ws = wb.create_sheet(title=sheet_name, index=0)
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, rows, header_info)

    if not date_groups:
        ws = wb.create_sheet(title="数据")
        _write_sheet(ws, [], header_info)

    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath


def export_excel(rows: list[dict], filename: str,
                 sheet_name: str = "送货单",
                 header_info: dict = None) -> str:
    """兼容旧接口 — 单Sheet导出"""
    return export_excel_single_sheet(rows, filename, sheet_name, header_info)
